import asyncio
import aiofiles
import openpyxl
from pathlib import Path

from src.logger import AsyncLogger


# Пути к файлам конфигурации
CONFIG_DIR = Path("config")
ACCOUNTS_EXCEL_PATH = CONFIG_DIR / "accounts.xlsx"
BAD_DISCORD_TOKENS_FILE = CONFIG_DIR / "bad_discord_token.txt"
BAD_TWITTER_TOKENS_FILE = CONFIG_DIR / "bad_twitter_token.txt"

# Названия колонок в Excel файле
DISCORD_TOKEN_COLUMN = 'Discord Token'
TWITTER_TOKEN_COLUMN = 'Twitter Token'

file_operation_lock = asyncio.Lock()
logger = AsyncLogger()


async def load_bad_tokens_from_file(file_path: Path) -> set[str]:
    """
    Загружает список плохих токенов из текстового файла
    """
    if not file_path.exists():
        return set()
    
    try:
        async with aiofiles.open(file_path, 'r', encoding='utf-8') as file:
            content = await file.read()
            tokens = {line.strip() for line in content.splitlines() if line.strip()}
            return tokens
    except Exception as e:
        await logger.logger_msg(
            f"Error when reading a file {file_path}: {str(e)}", "error", "load_bad_tokens_from_file"
        )
        return set()


async def is_token_already_marked_as_bad(file_path: Path, token: str) -> bool:
    """
    Проверяет, помечен ли токен как плохой
    """
    if not token or not isinstance(token, str):
        return False
    
    # Всегда загружаем актуальные данные из файла
    bad_tokens = await load_bad_tokens_from_file(file_path)
    return token.strip() in bad_tokens


def create_column_mapping_from_excel_header(worksheet) -> dict[str, int]:
    """
    Создает словарь соответствия названий колонок и их индексов
    """
    try:
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=False))
        column_mapping = {}
        
        for column_index, cell in enumerate(header_row):
            if cell.value is not None:
                column_name = str(cell.value).strip()
                if column_name:
                    column_mapping[column_name] = column_index
        
        return column_mapping
        
    except Exception as e:
        return {}


async def remove_token_from_excel_file(
        token: str, token_column_name: str, 
        wallet_address: str | None = None
    ) -> bool:
    """
    Удаляет указанный токен из Excel файла
    """
    if not token or not isinstance(token, str) or not token.strip():
        return False
    
    try:
        workbook = openpyxl.load_workbook(ACCOUNTS_EXCEL_PATH)
        worksheet = workbook.active
        column_mapping = create_column_mapping_from_excel_header(worksheet)
        
        token_column_index = column_mapping.get(token_column_name)
        
        rows_modified = 0
        token_to_find = token.strip()
        
        for row in worksheet.iter_rows(min_row=2, values_only=False):
            if token_column_index < len(row):
                cell = row[token_column_index]
                cell_value = cell.value
                if cell_value is not None:
                    cell_value_str = str(cell_value).strip()
                    if cell_value_str == token_to_find:
                        cell.value = ""
                        rows_modified += 1
        
        if rows_modified > 0:
            workbook.save(ACCOUNTS_EXCEL_PATH)
            await logger.logger_msg(
                f"Removed bad {token_column_name} from Excel file", "info", wallet_address
            )
            return True
        else:
            await logger.logger_msg(
                f"Token not found in Excel file for deletion", "warning", wallet_address, "remove_token_from_excel_file"
            )
            return False
            
    except Exception as e:
        await logger.logger_msg(
            f"Error when working with Excel file: {str(e)}", "error", wallet_address, "remove_token_from_excel_file"
        )
        return False


async def save_bad_token_to_file(file_path: Path, token: str) -> bool:
    """
    Сохраняет плохой токен в текстовый файл
    """
    if not token or not isinstance(token, str):
        return False
    
    token = token.strip()
    if not token:
        return False
    
    try:        
        if await is_token_already_marked_as_bad(file_path, token):
            return False
        
        async with aiofiles.open(file_path, 'a', encoding='utf-8') as file:
            await file.write(f"{token}\n")
        
        return True
        
    except Exception as e:
        return False


async def process_bad_token(
        token: str, token_column_name: str, bad_tokens_file: Path, 
        wallet_address: str | None = None
    ) -> None:
    """
    Универсальная функция обработки плохого токена
    """
    if not token or not isinstance(token, str):
        return
    
    try:
        token_saved = await save_bad_token_to_file(bad_tokens_file, token)
        token_removed = await remove_token_from_excel_file(
            token, token_column_name, wallet_address
        )
        
        if token_saved or token_removed:
            await logger.logger_msg(
                f"Bad token processed: saved = {token_saved}, deleted from Excel = {token_removed}", "info", wallet_address
            )
        
    except Exception as e:
        await logger.logger_msg(
            f"Error while processing a bad token: {str(e)}", "error", wallet_address, "process_bad_token"
        )


async def save_bad_discord_token(discord_token: str, wallet_address: str) -> None:
    """
    Обрабатывает недействительный Discord токен
    """
    async with file_operation_lock:
        await process_bad_token(
            token=discord_token,
            token_column_name=DISCORD_TOKEN_COLUMN,
            bad_tokens_file=BAD_DISCORD_TOKENS_FILE,
            wallet_address=wallet_address
        )


async def save_bad_twitter_token(twitter_token: str, wallet_address: str | None = None) -> None:
    """
    Обрабатывает недействительный Twitter токен
    """
    async with file_operation_lock:
        await process_bad_token(
            token=twitter_token,
            token_column_name=TWITTER_TOKEN_COLUMN,
            bad_tokens_file=BAD_TWITTER_TOKENS_FILE,
            wallet_address=wallet_address
        )